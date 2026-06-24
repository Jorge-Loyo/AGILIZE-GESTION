"""Servicio para importar empleados desde un archivo Excel."""
from pathlib import Path
from openpyxl import load_workbook
from core.database import get_db
from models.empleado import Empleado, Departamento, Cargo
from datetime import date


COLUMNAS_ESPERADAS = [
    "id", "legajo", "apellido", "nombre", "dni", "cuil", "email", "telefono",
    "direccion", "fecha_nacimiento", "fecha_ingreso",
    "departamento", "cargo", "valor_hora", "sueldo_mensual",
]


def importar_empleados(filepath: str) -> dict:
    """
    Importa empleados desde un Excel.
    Retorna dict con resultados: importados, errores, duplicados.
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Saltar header
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    resultados = {"importados": 0, "duplicados": 0, "errores": []}

    with get_db() as db:
        # Cache de departamentos y cargos
        deptos = {d.nombre.lower(): d.id for d in db.query(Departamento).all()}
        cargos = {c.nombre.lower(): c.id for c in db.query(Cargo).all()}

        for i, row in enumerate(rows, start=2):
            try:
                datos = {}
                for j, header in enumerate(headers):
                    if j < len(row):
                        datos[header] = row[j]

                # Flexibilidad: soportar nombre completo en un solo campo
                nombre_raw = str(datos.get("nombre", "") or "").strip()
                apellido_raw = str(datos.get("apellido", "") or "").strip()

                # Si apellido es un numero o vacio, intentar separar del nombre
                if not apellido_raw or apellido_raw.isdigit():
                    partes = nombre_raw.split(" ", 1)
                    if len(partes) == 2:
                        nombre = partes[0].strip()
                        apellido = partes[1].strip()
                    else:
                        nombre = nombre_raw
                        apellido = ""
                else:
                    nombre = nombre_raw
                    apellido = apellido_raw

                if not nombre:
                    resultados["errores"].append(f"Fila {i}: falta nombre")
                    continue

                # Legajo
                legajo_raw = str(datos.get("legajo", "") or "").strip()
                legajo = legajo_raw if legajo_raw else ""

                # DNI opcional
                dni = str(datos.get("dni", "") or "").strip().replace(".", "")

                # Verificar duplicado por legajo o DNI
                existente = None
                if legajo:
                    existente = db.query(Empleado).filter_by(legajo=legajo).first()
                if not existente and dni:
                    existente = db.query(Empleado).filter_by(dni=dni).first()

                if existente:
                    resultados["duplicados"] += 1
                    continue

                # Generar legajo si no tiene
                if not legajo:
                    max_count = db.query(Empleado).count() + resultados["importados"] + 1
                    legajo = str(max_count)

                cuil = str(datos.get("cuil", "") or "").strip()

                # Departamento
                depto_nombre = str(datos.get("departamento", "") or "").strip().lower()
                depto_id = None
                if depto_nombre:
                    if depto_nombre not in deptos:
                        nuevo_depto = Departamento(nombre=depto_nombre.title())
                        db.add(nuevo_depto)
                        db.flush()
                        deptos[depto_nombre] = nuevo_depto.id
                    depto_id = deptos[depto_nombre]

                # Cargo
                cargo_nombre = str(datos.get("cargo", "") or "").strip().lower()
                cargo_id = None
                if cargo_nombre:
                    if cargo_nombre not in cargos:
                        nuevo_cargo = Cargo(nombre=cargo_nombre.title())
                        db.add(nuevo_cargo)
                        db.flush()
                        cargos[cargo_nombre] = nuevo_cargo.id
                    cargo_id = cargos[cargo_nombre]

                # Fechas
                fecha_nac = _parse_fecha(datos.get("fecha_nacimiento"))
                fecha_ing = _parse_fecha(datos.get("fecha_ingreso")) or date.today()

                # Edad
                edad = None
                if fecha_nac:
                    hoy = date.today()
                    edad = hoy.year - fecha_nac.year - ((hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day))

                # Valores numéricos
                valor_hora = _parse_numero(datos.get("valor_hora"))
                sueldo_mensual = _parse_numero(datos.get("sueldo_mensual"))

                # Jornada por defecto desde config
                from services.core.empresa_service import empresa_service
                default_entrada = empresa_service.obtener("jornada_entrada") or "08:00"
                default_salida = empresa_service.obtener("jornada_salida") or "17:00"

                # Campos opcionales del Excel
                tipo_liq = str(datos.get("tipo_liquidacion", "") or "").strip().lower()
                if tipo_liq not in ("por_hora", "mensual"):
                    tipo_liq = "por_hora"
                hora_ent = str(datos.get("hora_entrada", "") or "").strip() or default_entrada
                hora_sal = str(datos.get("hora_salida", "") or "").strip() or default_salida
                dias_lab = str(datos.get("dias_laborales", "") or "").strip() or "lun,mar,mie,jue,vie"

                empleado = Empleado(
                    legajo=legajo,
                    apellido=apellido,
                    nombre=nombre,
                    dni=dni if dni else None,
                    cuil=cuil if cuil else None,
                    email=str(datos.get("email", "") or "").strip(),
                    telefono=str(datos.get("telefono", "") or "").strip(),
                    direccion=str(datos.get("direccion", "") or "").strip(),
                    fecha_nacimiento=fecha_nac,
                    edad=edad,
                    fecha_ingreso=fecha_ing,
                    departamento_id=depto_id,
                    cargo_id=cargo_id,
                    valor_hora=valor_hora or 0,
                    sueldo_mensual=sueldo_mensual or 0,
                    tipo_liquidacion=tipo_liq,
                    hora_entrada=hora_ent,
                    hora_salida=hora_sal,
                    dias_laborales=dias_lab,
                )
                db.add(empleado)
                resultados["importados"] += 1

            except Exception as e:
                resultados["errores"].append(f"Fila {i}: {str(e)[:80]}")

    wb.close()
    return resultados


def generar_plantilla(filepath: str):
    """Genera un Excel plantilla con los headers esperados."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    headers = ["Legajo", "Apellido", "Nombre", "DNI", "CUIL", "Email", "Telefono",
               "Direccion", "Fecha_Nacimiento", "Fecha_Ingreso",
               "Departamento", "Cargo", "Valor_Hora", "Sueldo_Mensual",
               "Tipo_Liquidacion", "Hora_Entrada", "Hora_Salida", "Dias_Laborales"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Ejemplos
    ws.append(["1", "Perez", "Juan", "12345678", "20-12345678-9", "juan@mail.com",
               "1155551234", "Calle 123", "15/03/1990", "01/06/2024",
               "Administracion", "Analista", "2500", "450000",
               "por_hora", "08:00", "17:00", "lun,mar,mie,jue,vie"])
    ws.append(["2", "Gomez", "Maria", "87654321", "27-87654321-3", "",
               "1166662345", "Av. Siempreviva 742", "20/08/1985", "15/03/2025",
               "Ventas", "Supervisora", "", "500000",
               "mensual", "09:00", "18:00", "lun,mar,mie,jue,vie"])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    wb.save(filepath)


def _parse_fecha(valor) -> date | None:
    if not valor:
        return None
    if isinstance(valor, date):
        return valor
    try:
        from datetime import datetime
        s = str(valor).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    except Exception:
        pass
    return None


def _parse_numero(valor) -> float | None:
    if not valor:
        return None
    try:
        return float(str(valor).replace(",", ".").replace("$", "").replace(" ", ""))
    except (ValueError, TypeError):
        return None
