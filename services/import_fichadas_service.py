"""
Importacion de fichadas desde archivos Excel.
Vincula por LEGAJO (campo "No" en XLS del reloj).
"""
from datetime import date, time, datetime
from pathlib import Path
from core.database import get_db
from models.empleado import Empleado
from services.asistencia_service import asistencia_service


def importar_fichadas(filepath: str, mapeo: dict = None) -> dict:
    """Importa fichadas. Detecta formato por extension."""
    ext = Path(filepath).suffix.lower()
    if ext == ".xls":
        return _importar_xls(filepath)
    return _importar_xlsx(filepath, mapeo=mapeo)


def _importar_xlsx(filepath: str, mapeo: dict = None) -> dict:
    """
    Formato XLSX manual: una hoja por empleado.
    Vincula por nombre de hoja contra nombre del empleado en BD.
    Si hay duplicados de nombre, marca error.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    resultados = {"importados": 0, "errores": [], "no_encontrados": []}

    hojas_ignorar = {"imprimir", "vacaciones aguinaldo", "hoja1", "hoja2", "hoja3", "hoja4", "hoja5", "sn"}

    with get_db() as db:
        empleados = db.query(Empleado).filter(Empleado.activo == True).all()

        nombre_map = {}
        for emp in empleados:
            key = emp.nombre.upper().strip()
            if key in nombre_map:
                nombre_map[key] = None
            else:
                nombre_map[key] = emp

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in hojas_ignorar:
                continue

            ws = wb[sheet_name]
            nombre_hoja = sheet_name.strip().upper()

            empleado = nombre_map.get(nombre_hoja)
            if empleado is None and nombre_hoja in nombre_map:
                resultados["errores"].append(f"{sheet_name}: nombre duplicado en BD, no se puede vincular sin legajo")
                continue
            if not empleado:
                # Intentar con mapeo manual
                if mapeo and sheet_name in mapeo:
                    emp_id = mapeo[sheet_name]
                    empleado = next((e for e in empleados if e.id == emp_id), None)
                if not empleado:
                    resultados["no_encontrados"].append(sheet_name)
                    continue

            for row in ws.iter_rows(min_row=8, max_col=4, values_only=True):
                try:
                    _, fecha_val, entrada_val, salida_val = row
                    if not fecha_val or not entrada_val or not salida_val:
                        continue

                    fecha = _parse_date(fecha_val)
                    hora_entrada = _parse_time(entrada_val)
                    hora_salida = _parse_time(salida_val)

                    if not fecha or not hora_entrada or not hora_salida:
                        continue

                    try:
                        asistencia_service.registrar(empleado.id, fecha, hora_entrada, hora_salida)
                        resultados["importados"] += 1
                    except ValueError:
                        pass
                except Exception:
                    pass

    wb.close()
    return resultados


def _importar_xls(filepath: str) -> dict:
    """
    Formato XLS del reloj fichador.
    Cada hoja tiene hasta 3 empleados en bloques de 15 columnas.
    Fila 2 col 9/24/39: Nombre (precedido por "Nom." en col 8/23/38)
    Fila 3 col 9/24/39: Legajo No
    Fichadas desde fila 11:
      - Col fecha: bloque_start (0/15/30)
      - Col entrada AM: bloque_start + 1
      - Col salida AM: bloque_start + 3
      - Col salida PM: bloque_start + 8
    """
    import xlrd

    wb = xlrd.open_workbook(filepath)
    resultados = {"importados": 0, "errores": [], "no_encontrados": []}

    # Leer periodo de la hoja Resumen (fila 1, col 1: "2026/05/16 ~ 05/31")
    ws_resumen = wb.sheet_by_index(0)
    periodo_raw = ""
    for c in range(ws_resumen.ncols):
        val = str(ws_resumen.cell_value(1, c)).strip()
        if "/" in val and "~" in val:
            periodo_raw = val
            break
    anio = None
    mes_inicio = None
    if "/" in periodo_raw:
        try:
            parts = periodo_raw.split("~")[0].strip().split("/")
            anio = int(parts[0])
            mes_inicio = int(parts[1])
        except (ValueError, IndexError):
            pass

    if not anio or not mes_inicio:
        resultados["errores"].append("No se pudo leer el periodo del archivo")
        return resultados

    with get_db() as db:
        empleados = db.query(Empleado).filter(Empleado.activo == True).all()
        emp_by_legajo = {emp.legajo: emp for emp in empleados if emp.legajo}

        # Bloques de empleado: offset de columnas
        bloques = [
            {"legajo_col": 9, "fecha_col": 0, "entrada_col": 1, "salida_cols": [8, 6, 3]},
            {"legajo_col": 24, "fecha_col": 15, "entrada_col": 16, "salida_cols": [23, 21, 18]},
            {"legajo_col": 39, "fecha_col": 30, "entrada_col": 31, "salida_cols": [38, 36, 33]},
        ]

        # Recorrer hojas de detalle (desde indice 4)
        for sheet_idx in range(4, wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)

            for bloque in bloques:
                if bloque["legajo_col"] >= ws.ncols:
                    continue

                # Leer legajo de fila 3
                legajo_raw = str(ws.cell_value(3, bloque["legajo_col"])).strip()
                if not legajo_raw or legajo_raw == "No":
                    continue

                # Convertir float a int string
                try:
                    legajo = str(int(float(legajo_raw)))
                except (ValueError, TypeError):
                    legajo = legajo_raw

                empleado = emp_by_legajo.get(legajo)
                if not empleado:
                    resultados["no_encontrados"].append(f"Legajo {legajo} (hoja {ws.name})")
                    continue

                # Leer fichadas desde fila 11
                for r in range(11, ws.nrows):
                    try:
                        fecha_str = str(ws.cell_value(r, bloque["fecha_col"])).strip()
                        if not fecha_str or len(fecha_str) < 4:
                            continue

                        dia_str = fecha_str.split(" ")[0]
                        if not dia_str.isdigit():
                            continue
                        dia = int(dia_str)

                        fecha = date(anio, mes_inicio, dia)

                        # Entrada AM
                        entrada_raw = str(ws.cell_value(r, bloque["entrada_col"])).strip()

                        # Salida: buscar en orden de prioridad
                        salida_raw = ""
                        for sal_col in bloque["salida_cols"]:
                            val = str(ws.cell_value(r, sal_col)).strip()
                            if val and val.lower() != "ausencia" and ":" in val:
                                salida_raw = val
                                break

                        if not entrada_raw or ":" not in entrada_raw:
                            continue
                        if "ausencia" in entrada_raw.lower():
                            continue

                        hora_entrada = _parse_time(entrada_raw)

                        # Si no hay salida, registrar con entrada = salida (marcado como incompleto)
                        if not salida_raw:
                            if hora_entrada:
                                try:
                                    asistencia_service.registrar(empleado.id, fecha, hora_entrada, hora_entrada, incompleto=True)
                                    resultados["importados"] += 1
                                    resultados["errores"].append(f"Leg {legajo} dia {dia}: entrada sin salida (incompleto)")
                                except ValueError:
                                    pass
                            continue

                        hora_salida = _parse_time(salida_raw)

                        if not hora_entrada or not hora_salida:
                            continue

                        try:
                            asistencia_service.registrar(empleado.id, fecha, hora_entrada, hora_salida)
                            resultados["importados"] += 1
                        except ValueError:
                            pass

                    except Exception:
                        pass

    return resultados


def _parse_date(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _parse_time(val) -> time | None:
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, (int, float)):
        hours = int(val * 24)
        minutes = int((val * 24 - hours) * 60)
        return time(hours % 24, minutes)
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).time()
            except ValueError:
                continue
    return None


def pre_scan_xlsx(filepath: str) -> list[str]:
    """Escanea un XLSX y retorna lista de hojas no vinculadas a empleados."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    hojas_ignorar = {"imprimir", "vacaciones aguinaldo", "hoja1", "hoja2", "hoja3", "hoja4", "hoja5", "sn"}
    no_encontrados = []

    with get_db() as db:
        empleados = db.query(Empleado).filter(Empleado.activo == True).all()
        nombre_map = {}
        for emp in empleados:
            key = emp.nombre.upper().strip()
            if key in nombre_map:
                nombre_map[key] = None
            else:
                nombre_map[key] = emp

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in hojas_ignorar:
                continue
            nombre_hoja = sheet_name.strip().upper()
            empleado = nombre_map.get(nombre_hoja)
            if not empleado:
                no_encontrados.append(sheet_name)

    wb.close()
    return no_encontrados
