from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from core.config import BASE_DIR

EXPORTS_DIR = BASE_DIR / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)


def exportar_excel(filename: str, headers: list[str], rows: list[list], titulo: str = "") -> str:
    """Exporta datos a un archivo Excel con formato. Retorna la ruta."""
    filepath = str(EXPORTS_DIR / filename)
    wb = Workbook()
    ws = wb.active
    ws.title = titulo or "Datos"

    # Estilo header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")

    # Título
    if titulo:
        ws.append([titulo])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(1, 1).font = Font(bold=True, size=14, color="D4AF37")
        ws.append([])

    # Headers
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(header_row, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row in rows:
        ws.append(row)

    # Auto-width
    for col_cells in ws.columns:
        col_letter = None
        max_length = 0
        for cell in col_cells:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    wb.save(filepath)
    return filepath
