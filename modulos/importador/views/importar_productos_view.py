"""
Importar Productos Masivamente desde Excel.
Permite cargar, previsualizar, mapear columnas y confirmar la importacion.
"""
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
    QMessageBox, QFrame, QComboBox, QGroupBox, QSpinBox,
)
from PySide6.QtCore import Qt
import qtawesome as qta


class ImportarProductosView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._df = None
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(12)

        title = QLabel("Importar Productos")
        title.setObjectName("title")
        layout.addWidget(title)

        subtitle = QLabel("Carga masiva de productos desde archivo Excel. Mapea las columnas y confirma.")
        subtitle.setStyleSheet("font-size: 11px; color: #888;")
        layout.addWidget(subtitle)

        # Boton cargar
        load_row = QHBoxLayout()
        btn_cargar = QPushButton("  Seleccionar Archivo Excel")
        btn_cargar.setIcon(qta.icon("fa5s.file-excel", color="#10b981"))
        btn_cargar.setFixedHeight(34)
        btn_cargar.setCursor(Qt.PointingHandCursor)
        btn_cargar.clicked.connect(self._cargar_archivo)
        load_row.addWidget(btn_cargar)

        btn_plantilla = QPushButton("  Descargar Plantilla")
        btn_plantilla.setIcon(qta.icon("fa5s.download", color="#3b82f6"))
        btn_plantilla.setFixedHeight(34)
        btn_plantilla.setCursor(Qt.PointingHandCursor)
        btn_plantilla.clicked.connect(self._descargar_plantilla)
        load_row.addWidget(btn_plantilla)

        self._lbl_archivo = QLabel("Ningun archivo seleccionado")
        self._lbl_archivo.setStyleSheet("font-size: 11px; color: #888;")
        load_row.addWidget(self._lbl_archivo)
        load_row.addStretch()
        layout.addLayout(load_row)

        # Mapeo de columnas
        grp_mapeo = QGroupBox("Mapeo de Columnas")
        grp_mapeo.setStyleSheet("QGroupBox { font-weight: bold; font-size: 11px; padding-top: 12px; }")
        mapeo_lay = QHBoxLayout(grp_mapeo)
        mapeo_lay.setSpacing(8)

        self._mapeos = {}
        campos = [
            ("Codigo:", "codigo"),
            ("Nombre:", "nombre"),
            ("Costo:", "costo"),
            ("P. Venta:", "precio_venta"),
            ("Stock:", "stock"),
            ("Categoria:", "categoria"),
        ]
        for label, key in campos:
            mapeo_lay.addWidget(QLabel(label))
            combo = QComboBox()
            combo.setFixedHeight(26)
            combo.setMinimumWidth(80)
            combo.addItem("-- No importar --", -1)
            mapeo_lay.addWidget(combo)
            self._mapeos[key] = combo

        layout.addWidget(grp_mapeo)

        # Fila inicio
        inicio_row = QHBoxLayout()
        inicio_row.addWidget(QLabel("Fila de inicio de datos:"))
        self._spin_inicio = QSpinBox()
        self._spin_inicio.setRange(1, 100)
        self._spin_inicio.setValue(2)
        self._spin_inicio.setFixedHeight(26)
        self._spin_inicio.setFixedWidth(60)
        inicio_row.addWidget(self._spin_inicio)
        inicio_row.addStretch()

        self._lbl_resumen = QLabel("")
        self._lbl_resumen.setStyleSheet("font-size: 12px; font-weight: bold; color: #D4AF37;")
        inicio_row.addWidget(self._lbl_resumen)
        layout.addLayout(inicio_row)

        # Preview
        self._tabla = QTableWidget()
        self._tabla.setAlternatingRowColors(True)
        self._tabla.verticalHeader().setVisible(False)
        self._tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tabla.setMaximumHeight(200)
        layout.addWidget(self._tabla)

        # Boton importar
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._btn_importar = QPushButton("  Importar Productos")
        self._btn_importar.setIcon(qta.icon("fa5s.database", color="#0f0f0f"))
        self._btn_importar.setFixedHeight(36)
        self._btn_importar.setFixedWidth(200)
        self._btn_importar.setCursor(Qt.PointingHandCursor)
        self._btn_importar.setEnabled(False)
        self._btn_importar.clicked.connect(self._importar)
        btn_row.addWidget(self._btn_importar)
        layout.addLayout(btn_row)

        layout.addStretch()

    def _cargar_archivo(self):
        ruta, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo", "", "Excel (*.xls *.xlsx)"
        )
        if not ruta:
            return

        try:
            import pandas as pd
            engine = "xlrd" if ruta.endswith(".xls") else "openpyxl"
            self._df = pd.read_excel(ruta, engine=engine, header=None)
            self._lbl_archivo.setText(ruta.split("/")[-1].split("\\")[-1])

            # Llenar combos con columnas
            ncols = len(self._df.columns)
            for key, combo in self._mapeos.items():
                combo.clear()
                combo.addItem("-- No importar --", -1)
                for col in range(ncols):
                    # Mostrar valor de la primera fila como referencia
                    sample = str(self._df.iloc[0, col])[:20] if len(self._df) > 0 else f"Col {col}"
                    combo.addItem(f"Col {col}: {sample}", col)

            # Auto-mapear por posicion comun
            defaults = {"codigo": 0, "nombre": 1, "costo": 2, "precio_venta": 3, "stock": 4}
            for key, col_idx in defaults.items():
                if key in self._mapeos and col_idx < ncols:
                    self._mapeos[key].setCurrentIndex(col_idx + 1)

            # Preview
            self._mostrar_preview()
            self._btn_importar.setEnabled(True)
            self._lbl_resumen.setText(f"{len(self._df)} filas encontradas")

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _mostrar_preview(self):
        if self._df is None:
            return
        preview = self._df.head(10)
        self._tabla.setColumnCount(len(preview.columns))
        self._tabla.setRowCount(len(preview))
        self._tabla.setHorizontalHeaderLabels([f"Col {i}" for i in range(len(preview.columns))])
        for i in range(len(preview)):
            for j in range(len(preview.columns)):
                val = str(preview.iloc[i, j]) if not __import__('pandas').isna(preview.iloc[i, j]) else ""
                self._tabla.setItem(i, j, QTableWidgetItem(val[:30]))
        self._tabla.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def _importar(self):
        if self._df is None:
            return

        col_codigo = self._mapeos["codigo"].currentData()
        col_nombre = self._mapeos["nombre"].currentData()

        if col_codigo == -1 or col_nombre == -1:
            QMessageBox.warning(self, "Error", "Codigo y Nombre son obligatorios para importar.")
            return

        col_costo = self._mapeos["costo"].currentData()
        col_venta = self._mapeos["precio_venta"].currentData()
        col_stock = self._mapeos["stock"].currentData()
        col_cat = self._mapeos["categoria"].currentData()
        fila_inicio = self._spin_inicio.value() - 1

        import pandas as pd
        from services.inventario import inventario_service

        creados = 0
        actualizados = 0
        errores = 0

        for i in range(fila_inicio, len(self._df)):
            row = self._df.iloc[i]
            try:
                codigo = str(row.iloc[col_codigo]).strip() if pd.notna(row.iloc[col_codigo]) else ""
                nombre = str(row.iloc[col_nombre]).strip() if pd.notna(row.iloc[col_nombre]) else ""

                if not codigo or not nombre:
                    continue

                datos = {"codigo": codigo, "nombre": nombre}

                if col_costo != -1 and pd.notna(row.iloc[col_costo]):
                    datos["precio_costo"] = float(row.iloc[col_costo])
                if col_venta != -1 and pd.notna(row.iloc[col_venta]):
                    datos["precio_venta"] = float(row.iloc[col_venta])
                if col_stock != -1 and pd.notna(row.iloc[col_stock]):
                    datos["stock_minimo"] = int(float(row.iloc[col_stock]))

                # Verificar si existe
                existentes = inventario_service.buscar_productos(codigo)
                existe = next((p for p in existentes if p.codigo == codigo), None)

                if existe:
                    inventario_service.actualizar_producto(existe.id, datos)
                    actualizados += 1
                else:
                    inventario_service.crear_producto(datos)
                    creados += 1

            except Exception:
                errores += 1

        QMessageBox.information(
            self, "Importacion Completada",
            f"Resultados:\n\n"
            f"Creados: {creados}\n"
            f"Actualizados: {actualizados}\n"
            f"Errores: {errores}\n"
            f"Total procesados: {creados + actualizados + errores}"
        )

    def _descargar_plantilla(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla", "plantilla_productos.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos"
            headers = ["Codigo", "Nombre", "Precio Costo", "Precio Venta", "Stock Minimo", "Categoria"]
            fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = fill
            # Ejemplo
            ws.append(["PROD-001", "Producto ejemplo", 10.00, 15.00, 5, "General"])
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 18
            wb.save(path)
            QMessageBox.information(self, "Exito", f"Plantilla guardada en:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
