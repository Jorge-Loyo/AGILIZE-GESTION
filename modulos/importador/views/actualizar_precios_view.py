"""
Actualizar Precios Masivamente desde Excel.
Busca productos por codigo y actualiza precio costo y/o venta.
"""
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
    QMessageBox, QGroupBox, QComboBox, QSpinBox, QCheckBox,
)
from PySide6.QtCore import Qt
import qtawesome as qta


class ActualizarPreciosView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._df = None
        self._resultados = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(12)

        title = QLabel("Actualizar Precios")
        title.setObjectName("title")
        layout.addWidget(title)

        subtitle = QLabel(
            "Carga un Excel con codigos de producto y nuevos precios.\n"
            "Los productos se buscan por codigo y se actualizan automaticamente."
        )
        subtitle.setStyleSheet("font-size: 11px; color: #888;")
        layout.addWidget(subtitle)

        # Cargar
        load_row = QHBoxLayout()
        btn_cargar = QPushButton("  Seleccionar Archivo")
        btn_cargar.setIcon(qta.icon("fa5s.file-excel", color="#10b981"))
        btn_cargar.setFixedHeight(34)
        btn_cargar.setCursor(Qt.PointingHandCursor)
        btn_cargar.clicked.connect(self._cargar_archivo)
        load_row.addWidget(btn_cargar)

        btn_plantilla = QPushButton("  Descargar Plantilla")
        btn_plantilla.setIcon(qta.icon("fa5s.download", color="#3b82f6"))
        btn_plantilla.setFixedHeight(34)
        btn_plantilla.setCursor(Qt.PointingHandCursor)
        btn_plantilla.clicked.connect(self._descargar_plantilla)
        load_row.addWidget(btn_plantilla)

        self._lbl_archivo = QLabel("Ningun archivo seleccionado")
        self._lbl_archivo.setStyleSheet("font-size: 11px; color: #888;")
        load_row.addWidget(self._lbl_archivo)
        load_row.addStretch()
        layout.addLayout(load_row)

        # Mapeo
        grp_mapeo = QGroupBox("Mapeo de Columnas")
        grp_mapeo.setStyleSheet("QGroupBox { font-weight: bold; font-size: 11px; padding-top: 12px; }")
        mapeo_lay = QHBoxLayout(grp_mapeo)
        mapeo_lay.setSpacing(8)

        self._mapeos = {}
        campos = [
            ("Codigo:", "codigo"),
            ("P. Costo:", "precio_costo"),
            ("P. Venta:", "precio_venta"),
        ]
        for label, key in campos:
            mapeo_lay.addWidget(QLabel(label))
            combo = QComboBox()
            combo.setFixedHeight(26)
            combo.setMinimumWidth(100)
            combo.addItem("-- No usar --", -1)
            mapeo_lay.addWidget(combo)
            self._mapeos[key] = combo

        mapeo_lay.addStretch()

        mapeo_lay.addWidget(QLabel("Fila inicio:"))
        self._spin_inicio = QSpinBox()
        self._spin_inicio.setRange(1, 100)
        self._spin_inicio.setValue(2)
        self._spin_inicio.setFixedHeight(26)
        self._spin_inicio.setFixedWidth(50)
        mapeo_lay.addWidget(self._spin_inicio)

        layout.addWidget(grp_mapeo)

        # Opciones
        opts_row = QHBoxLayout()
        self._chk_solo_existentes = QCheckBox("Solo actualizar existentes (no crear nuevos)")
        self._chk_solo_existentes.setChecked(True)
        opts_row.addWidget(self._chk_solo_existentes)
        opts_row.addStretch()
        layout.addLayout(opts_row)

        # Botones
        btn_row = QHBoxLayout()
        self._btn_preview = QPushButton("  Previsualizar Cambios")
        self._btn_preview.setIcon(qta.icon("fa5s.eye", color="#3b82f6"))
        self._btn_preview.setFixedHeight(32)
        self._btn_preview.setEnabled(False)
        self._btn_preview.clicked.connect(self._previsualizar)
        btn_row.addWidget(self._btn_preview)

        btn_row.addStretch()

        self._lbl_resumen = QLabel("")
        self._lbl_resumen.setStyleSheet("font-size: 12px; font-weight: bold; color: #D4AF37;")
        btn_row.addWidget(self._lbl_resumen)

        self._btn_aplicar = QPushButton("  Aplicar Cambios")
        self._btn_aplicar.setIcon(qta.icon("fa5s.check", color="#0f0f0f"))
        self._btn_aplicar.setFixedHeight(32)
        self._btn_aplicar.setFixedWidth(180)
        self._btn_aplicar.setEnabled(False)
        self._btn_aplicar.clicked.connect(self._aplicar)
        btn_row.addWidget(self._btn_aplicar)
        layout.addLayout(btn_row)

        # Tabla preview
        self._tabla = QTableWidget()
        self._tabla.setColumnCount(6)
        self._tabla.setHorizontalHeaderLabels([
            "Codigo", "Producto", "Costo Anterior", "Costo Nuevo", "Venta Anterior", "Venta Nuevo"
        ])
        self._tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self._tabla.setAlternatingRowColors(True)
        self._tabla.verticalHeader().setVisible(False)
        self._tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self._tabla, 1)

    def _cargar_archivo(self):
        ruta, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo de precios", "", "Excel (*.xls *.xlsx);;CSV (*.csv)"
        )
        if not ruta:
            return

        try:
            import pandas as pd
            if ruta.endswith(".csv"):
                self._df = pd.read_csv(ruta, header=None)
            else:
                engine = "xlrd" if ruta.endswith(".xls") else "openpyxl"
                self._df = pd.read_excel(ruta, engine=engine, header=None)

            self._lbl_archivo.setText(ruta.split("/")[-1].split("\\")[-1])

            ncols = len(self._df.columns)
            for key, combo in self._mapeos.items():
                combo.clear()
                combo.addItem("-- No usar --", -1)
                for col in range(ncols):
                    sample = str(self._df.iloc[0, col])[:20] if len(self._df) > 0 else f"Col {col}"
                    combo.addItem(f"Col {col}: {sample}", col)

            # Auto-mapear
            if ncols >= 1:
                self._mapeos["codigo"].setCurrentIndex(1)
            if ncols >= 2:
                self._mapeos["precio_costo"].setCurrentIndex(2)
            if ncols >= 3:
                self._mapeos["precio_venta"].setCurrentIndex(3)

            self._btn_preview.setEnabled(True)
            self._lbl_resumen.setText(f"{len(self._df)} filas en archivo")

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _previsualizar(self):
        if self._df is None:
            return

        col_codigo = self._mapeos["codigo"].currentData()
        col_costo = self._mapeos["precio_costo"].currentData()
        col_venta = self._mapeos["precio_venta"].currentData()

        if col_codigo == -1:
            QMessageBox.warning(self, "Error", "La columna de codigo es obligatoria.")
            return

        if col_costo == -1 and col_venta == -1:
            QMessageBox.warning(self, "Error", "Selecciona al menos una columna de precio.")
            return

        import pandas as pd
        from services.inventario_service import inventario_service

        fila_inicio = self._spin_inicio.value() - 1
        self._resultados = []

        for i in range(fila_inicio, len(self._df)):
            row = self._df.iloc[i]
            codigo = str(row.iloc[col_codigo]).strip() if pd.notna(row.iloc[col_codigo]) else ""
            if not codigo:
                continue

            existentes = inventario_service.buscar_productos(codigo)
            producto = next((p for p in existentes if p.codigo == codigo), None)

            if not producto and self._chk_solo_existentes.isChecked():
                continue

            nuevo_costo = None
            nuevo_venta = None

            if col_costo != -1 and pd.notna(row.iloc[col_costo]):
                try:
                    nuevo_costo = float(row.iloc[col_costo])
                except (ValueError, TypeError):
                    pass

            if col_venta != -1 and pd.notna(row.iloc[col_venta]):
                try:
                    nuevo_venta = float(row.iloc[col_venta])
                except (ValueError, TypeError):
                    pass

            self._resultados.append({
                "codigo": codigo,
                "nombre": producto.nombre if producto else "(NUEVO)",
                "costo_anterior": producto.precio_costo if producto else 0,
                "costo_nuevo": nuevo_costo,
                "venta_anterior": producto.precio_venta if producto else 0,
                "venta_nuevo": nuevo_venta,
                "producto_id": producto.id if producto else None,
            })

        # Mostrar en tabla
        self._tabla.setRowCount(len(self._resultados))
        for i, r in enumerate(self._resultados):
            self._tabla.setItem(i, 0, QTableWidgetItem(r["codigo"]))
            self._tabla.setItem(i, 1, QTableWidgetItem(r["nombre"]))
            self._tabla.setItem(i, 2, QTableWidgetItem(f"$ {r['costo_anterior']:,.2f}"))
            self._tabla.setItem(i, 3, QTableWidgetItem(f"$ {r['costo_nuevo']:,.2f}" if r["costo_nuevo"] is not None else "---"))
            self._tabla.setItem(i, 4, QTableWidgetItem(f"$ {r['venta_anterior']:,.2f}"))
            self._tabla.setItem(i, 5, QTableWidgetItem(f"$ {r['venta_nuevo']:,.2f}" if r["venta_nuevo"] is not None else "---"))

        self._lbl_resumen.setText(f"{len(self._resultados)} productos a actualizar")
        self._btn_aplicar.setEnabled(len(self._resultados) > 0)

    def _aplicar(self):
        if not self._resultados:
            return

        resp = QMessageBox.question(
            self, "Confirmar",
            f"Se actualizaran {len(self._resultados)} productos.\nContinuar?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        from services.inventario_service import inventario_service

        actualizados = 0
        errores = 0

        for r in self._resultados:
            try:
                if r["producto_id"]:
                    datos = {}
                    if r["costo_nuevo"] is not None:
                        datos["precio_costo"] = r["costo_nuevo"]
                    if r["venta_nuevo"] is not None:
                        datos["precio_venta"] = r["venta_nuevo"]
                    if datos:
                        inventario_service.actualizar_producto(r["producto_id"], datos)
                        actualizados += 1
            except Exception:
                errores += 1

        QMessageBox.information(
            self, "Actualizacion Completada",
            f"Resultados:\n\n"
            f"Actualizados: {actualizados}\n"
            f"Errores: {errores}"
        )
        self._btn_aplicar.setEnabled(False)

    def _descargar_plantilla(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla", "plantilla_precios.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Precios"
            headers = ["Codigo", "Precio Costo", "Precio Venta"]
            fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = fill
            ws.append(["PROD-001", 10.00, 15.00])
            for col in range(1, 4):
                ws.column_dimensions[chr(64 + col)].width = 18
            wb.save(path)
            QMessageBox.information(self, "Exito", f"Plantilla guardada en:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
