"""Vista de Listas de Precios de Proveedores."""
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QMessageBox, QDialog, QLineEdit, QFileDialog, QSpinBox,
)
from PySide6.QtCore import Qt
import qtawesome as qta


class PreciosProveedoresView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self._cargar()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        # Header
        header = QHBoxLayout()
        title = QLabel("Listas de Precios - Proveedores")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #D4AF37;")
        header.addWidget(title)
        header.addStretch()

        btn_nueva = QPushButton("  Nueva Lista")
        btn_nueva.setIcon(qta.icon("fa5s.plus", color="#0f0f0f"))
        btn_nueva.setFixedHeight(30)
        btn_nueva.clicked.connect(self._nueva_lista)
        header.addWidget(btn_nueva)

        btn_importar = QPushButton("  Importar Excel")
        btn_importar.setIcon(qta.icon("fa5s.file-excel", color="#0f0f0f"))
        btn_importar.setFixedHeight(30)
        btn_importar.clicked.connect(self._importar_excel)
        header.addWidget(btn_importar)

        btn_plantilla = QPushButton("  Plantilla")
        btn_plantilla.setIcon(qta.icon("fa5s.download", color="#3b82f6"))
        btn_plantilla.setFixedHeight(30)
        btn_plantilla.clicked.connect(self._descargar_plantilla)
        header.addWidget(btn_plantilla)
        layout.addLayout(header)

        # Filtro proveedor
        filtro = QHBoxLayout()
        filtro.addWidget(QLabel("Proveedor:"))
        self._combo_prov = QComboBox()
        self._combo_prov.setFixedHeight(28)
        self._combo_prov.setMinimumWidth(200)
        self._combo_prov.addItem("-- Todos --", None)
        try:
            from core.database import get_db
            from models.datos import Proveedor
            with get_db() as db:
                for p in db.query(Proveedor).filter(Proveedor.activo == True).order_by(Proveedor.razon_social).all():
                    self._combo_prov.addItem(p.razon_social, p.id)
        except Exception:
            pass
        self._combo_prov.currentIndexChanged.connect(self._cargar)
        filtro.addWidget(self._combo_prov)
        filtro.addStretch()
        layout.addLayout(filtro)

        # Tabla listas
        self._tabla = QTableWidget()
        self._tabla.setColumnCount(6)
        self._tabla.setHorizontalHeaderLabels(["ID", "Proveedor", "Nombre Lista", "Fecha", "Moneda", "Vigente"])
        self._tabla.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self._tabla.setAlternatingRowColors(True)
        self._tabla.verticalHeader().setVisible(False)
        self._tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tabla.setSelectionBehavior(QTableWidget.SelectRows)
        self._tabla.setColumnHidden(0, True)
        self._tabla.doubleClicked.connect(self._ver_detalle)
        layout.addWidget(self._tabla)

        # Botones abajo
        bot = QHBoxLayout()
        btn_ver = QPushButton("  Ver Detalle")
        btn_ver.setIcon(qta.icon("fa5s.eye", color="#0f0f0f"))
        btn_ver.setFixedHeight(30)
        btn_ver.clicked.connect(self._ver_detalle)
        bot.addWidget(btn_ver)

        btn_eliminar = QPushButton("  Eliminar")
        btn_eliminar.setIcon(qta.icon("fa5s.trash", color="#ef4444"))
        btn_eliminar.setFixedHeight(30)
        btn_eliminar.clicked.connect(self._eliminar)
        bot.addWidget(btn_eliminar)
        bot.addStretch()
        layout.addLayout(bot)

    def _cargar(self):
        from services.compras.compras_service import compras_service
        prov_id = self._combo_prov.currentData()
        listas = compras_service.listar_listas_precio(proveedor_id=prov_id)
        self._tabla.setRowCount(len(listas))
        for i, l in enumerate(listas):
            self._tabla.setItem(i, 0, QTableWidgetItem(str(l.id)))
            prov_nombre = l.proveedor.razon_social if l.proveedor else ""
            self._tabla.setItem(i, 1, QTableWidgetItem(prov_nombre))
            self._tabla.setItem(i, 2, QTableWidgetItem(l.nombre))
            self._tabla.setItem(i, 3, QTableWidgetItem(l.fecha.strftime("%d/%m/%Y") if l.fecha else ""))
            self._tabla.setItem(i, 4, QTableWidgetItem(l.moneda))
            vig = QTableWidgetItem("Si" if l.vigente else "No")
            vig.setTextAlignment(Qt.AlignCenter)
            self._tabla.setItem(i, 5, vig)

    def _nueva_lista(self):
        dlg = _NuevaListaDialog(self)
        if dlg.exec() == QDialog.Accepted:
            data = dlg.datos()
            from services.compras.compras_service import compras_service
            compras_service.crear_lista_precio(
                proveedor_id=data["proveedor_id"],
                nombre=data["nombre"],
                moneda=data["moneda"],
                items=data["items"],
            )
            self._cargar()

    def _importar_excel(self):
        """Importar lista de precios desde Excel. Espera columnas: codigo/descripcion/precio/descuento."""
        # Seleccionar proveedor
        prov_id = self._combo_prov.currentData()
        if not prov_id:
            QMessageBox.warning(self, "Aviso", "Selecciona un proveedor primero en el filtro.")
            return

        path, _ = QFileDialog.getOpenFileName(self, "Seleccionar Excel", "", "Excel (*.xlsx *.xls)")
        if not path:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            items = []
            for row in rows:
                if not row or not row[0]:
                    continue
                items.append({
                    "codigo_proveedor": str(row[0]) if row[0] else "",
                    "descripcion": str(row[1]) if len(row) > 1 and row[1] else str(row[0]),
                    "precio_unitario": float(row[2]) if len(row) > 2 and row[2] else 0,
                    "descuento": float(row[3]) if len(row) > 3 and row[3] else 0,
                })
            wb.close()

            if not items:
                QMessageBox.warning(self, "Aviso", "No se encontraron items en el archivo.")
                return

            import os
            nombre_lista = os.path.splitext(os.path.basename(path))[0]
            from services.compras.compras_service import compras_service
            compras_service.crear_lista_precio(
                proveedor_id=prov_id, nombre=nombre_lista, items=items
            )
            QMessageBox.information(self, "Exito", f"Lista importada: {len(items)} items.")
            self._cargar()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al importar: {e}")

    def _ver_detalle(self):
        row = self._tabla.currentRow()
        if row < 0:
            return
        lista_id = int(self._tabla.item(row, 0).text())
        nombre = self._tabla.item(row, 2).text()
        dlg = _DetalleListaDialog(lista_id, nombre, self)
        dlg.exec()

    def _eliminar(self):
        row = self._tabla.currentRow()
        if row < 0:
            return
        lista_id = int(self._tabla.item(row, 0).text())
        if QMessageBox.question(self, "Confirmar", "Eliminar esta lista de precios?") == QMessageBox.Yes:
            from services.compras.compras_service import compras_service
            compras_service.eliminar_lista_precio(lista_id)
            self._cargar()

    def _descargar_plantilla(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla", "plantilla_lista_precios.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Lista Precios"
            headers = ["Codigo Proveedor", "Descripcion", "Precio Unitario", "Descuento %"]
            fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = fill
            ws.append(["PROV-001", "Producto ejemplo", 25.50, 10])
            ws.append(["PROV-002", "Otro producto", 12.00, 0])
            for col in range(1, 5):
                ws.column_dimensions[chr(64 + col)].width = 20
            wb.save(path)
            QMessageBox.information(self, "Exito", f"Plantilla guardada en:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


class _NuevaListaDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Nueva Lista de Precios")
        self.setMinimumSize(500, 300)
        self._items = []
        layout = QVBoxLayout(self)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Proveedor:"))
        self._combo_prov = QComboBox()
        self._combo_prov.setMinimumWidth(200)
        try:
            from core.database import get_db
            from models.datos import Proveedor
            with get_db() as db:
                for p in db.query(Proveedor).filter(Proveedor.activo == True).order_by(Proveedor.razon_social).all():
                    self._combo_prov.addItem(p.razon_social, p.id)
        except Exception:
            pass
        row1.addWidget(self._combo_prov)
        row1.addWidget(QLabel("Moneda:"))
        self._combo_moneda = QComboBox()
        self._combo_moneda.addItems(["USD", "ARS", "VES", "EUR"])
        row1.addWidget(self._combo_moneda)
        layout.addLayout(row1)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Nombre:"))
        self._input_nombre = QLineEdit()
        self._input_nombre.setMaxLength(200)
        self._input_nombre.setPlaceholderText("Ej: Lista Enero 2025")
        row2.addWidget(self._input_nombre)
        layout.addLayout(row2)

        # Items manuales
        lbl = QLabel("Items (opcional - tambien puedes importar desde Excel luego):")
        lbl.setStyleSheet("font-size: 11px; color: #888;")
        layout.addWidget(lbl)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ok = QPushButton("Crear Lista")
        btn_ok.setFixedHeight(32)
        btn_ok.clicked.connect(self._guardar)
        btn_row.addWidget(btn_ok)
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setFixedHeight(32)
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _guardar(self):
        if not self._combo_prov.currentData():
            QMessageBox.warning(self, "Error", "Selecciona un proveedor.")
            return
        if not self._input_nombre.text().strip():
            QMessageBox.warning(self, "Error", "Ingresa un nombre para la lista.")
            return
        self.accept()

    def datos(self):
        return {
            "proveedor_id": self._combo_prov.currentData(),
            "nombre": self._input_nombre.text().strip(),
            "moneda": self._combo_moneda.currentText(),
            "items": [],
        }


class _DetalleListaDialog(QDialog):
    def __init__(self, lista_id, nombre, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Detalle: {nombre}")
        self.setMinimumSize(700, 400)
        layout = QVBoxLayout(self)

        from services.compras.compras_service import compras_service
        detalles = compras_service.obtener_lista_detalles(lista_id)

        tabla = QTableWidget()
        tabla.setColumnCount(5)
        tabla.setHorizontalHeaderLabels(["Cod. Proveedor", "Descripcion", "Precio", "Descuento %", "Precio Neto"])
        tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        tabla.setAlternatingRowColors(True)
        tabla.verticalHeader().setVisible(False)
        tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        tabla.setRowCount(len(detalles))
        for i, d in enumerate(detalles):
            tabla.setItem(i, 0, QTableWidgetItem(d.codigo_proveedor))
            tabla.setItem(i, 1, QTableWidgetItem(d.descripcion))
            tabla.setItem(i, 2, QTableWidgetItem(f"$ {d.precio_unitario:,.2f}"))
            tabla.setItem(i, 3, QTableWidgetItem(f"{d.descuento}%"))
            tabla.setItem(i, 4, QTableWidgetItem(f"$ {d.precio_neto:,.2f}"))
        layout.addWidget(tabla)

        btn = QPushButton("Cerrar")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn, alignment=Qt.AlignRight)
