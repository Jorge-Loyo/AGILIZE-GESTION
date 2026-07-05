from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QStackedWidget,
    QMessageBox, QCheckBox, QPushButton, QFileDialog,
    QDialog, QLabel, QScrollArea, QFrame, QComboBox,
)
from PySide6.QtCore import Qt
import qtawesome as qta
from ui.components.data_table import DataTable
from modulos.rrhh.views.form_empleado import FormEmpleado
from modulos.rrhh.views.detalle_empleado_dialog import EmpleadoDetalleDialog
from services.rrhh.empleado_service import empleado_service
from services.herramientas.export_service import exportar_excel
from services.herramientas.import_service import importar_empleados, generar_plantilla
from services.core.pais_config_service import label_doc_identidad, label_id_fiscal
import os


class EmpleadosView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self._cargar_lista()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(0)

        self.stack = QStackedWidget()

        # Pagina 0: Lista
        lista_page = QWidget()
        lista_layout = QVBoxLayout(lista_page)
        lista_layout.setContentsMargins(0, 0, 0, 0)
        lista_layout.setSpacing(8)

        self.tabla = DataTable(["Legajo", "Apellido", "Nombre", label_doc_identidad(), label_id_fiscal(), "Departamento", "Cargo", "Estado"])
        self.tabla.btn_buscar.clicked.connect(self._cargar_lista)
        self.tabla.input_busqueda.returnPressed.connect(self._cargar_lista)
        self.tabla.btn_nuevo.clicked.connect(self._nuevo)
        self.tabla.row_double_clicked.connect(self._editar)
        lista_layout.addWidget(self.tabla)

        # Fila 1: Filtro + acciones sobre empleado
        row1 = QHBoxLayout()
        row1.setSpacing(8)

        self.chk_inactivos = QCheckBox("Mostrar inactivos")
        self.chk_inactivos.stateChanged.connect(self._cargar_lista)
        row1.addWidget(self.chk_inactivos)

        row1.addWidget(QLabel("Ordenar:"))
        self.filtro_orden = QComboBox()
        self.filtro_orden.setMinimumHeight(32)
        self.filtro_orden.addItem("Legajo", "legajo")
        self.filtro_orden.addItem("Nombre", "nombre")
        self.filtro_orden.addItem("Apellido", "apellido")
        self.filtro_orden.currentIndexChanged.connect(self._cargar_lista)
        row1.addWidget(self.filtro_orden)

        row1.addStretch()

        btn_ver = QPushButton("  Editar")
        btn_ver.setIcon(qta.icon("fa5s.edit", color="#ffffff"))
        btn_ver.setCursor(Qt.PointingHandCursor)
        btn_ver.setToolTip("Editar el empleado seleccionado")
        btn_ver.setStyleSheet("QPushButton { background-color: #2D2D2D; color: #F8F9FA; } QPushButton:hover { background-color: #3d3d3d; }")
        btn_ver.clicked.connect(self._editar_seleccionado)
        row1.addWidget(btn_ver)

        btn_baja = QPushButton("  Dar de baja")
        btn_baja.setIcon(qta.icon("fa5s.user-minus", color="#ffffff"))
        btn_baja.setCursor(Qt.PointingHandCursor)
        btn_baja.setToolTip("Desactivar el empleado seleccionado")
        btn_baja.setStyleSheet("QPushButton { background-color: #ef4444; } QPushButton:hover { background-color: #dc2626; }")
        btn_baja.clicked.connect(self._eliminar)
        row1.addWidget(btn_baja)

        lista_layout.addLayout(row1)

        # Fila 2: Herramientas (PDF / Excel)
        row2 = QHBoxLayout()
        row2.setSpacing(8)

        btn_formulario = QPushButton("  Formulario Alta (PDF)")
        btn_formulario.setIcon(qta.icon("fa5s.file-pdf", color="#121212"))
        btn_formulario.setCursor(Qt.PointingHandCursor)
        btn_formulario.setToolTip("Genera un PDF en blanco para que complete el nuevo empleado")
        btn_formulario.setStyleSheet("QPushButton { background-color: #D4AF37; color: #121212; font-weight: bold; } QPushButton:hover { background-color: #c9a030; }")
        btn_formulario.clicked.connect(self._generar_formulario_alta)
        row2.addWidget(btn_formulario)

        row2.addStretch()

        btn_export = QPushButton("  Exportar Excel")
        btn_export.setIcon(qta.icon("fa5s.file-export", color="#0f0f0f"))
        btn_export.setCursor(Qt.PointingHandCursor)
        btn_export.setToolTip("Exporta el listado de empleados activos e inactivos a Excel")
        btn_export.setStyleSheet("QPushButton { background-color: #10b981; } QPushButton:hover { background-color: #059669; }")
        btn_export.clicked.connect(self._exportar)
        row2.addWidget(btn_export)

        btn_importar = QPushButton("  Importar Excel")
        btn_importar.setIcon(qta.icon("fa5s.file-import", color="#ffffff"))
        btn_importar.setCursor(Qt.PointingHandCursor)
        btn_importar.setToolTip("Importa empleados desde un archivo Excel")
        btn_importar.setStyleSheet("QPushButton { background-color: #6366f1; } QPushButton:hover { background-color: #4f46e5; }")
        btn_importar.clicked.connect(self._importar)
        row2.addWidget(btn_importar)

        btn_plantilla = QPushButton("  Descargar Plantilla")
        btn_plantilla.setIcon(qta.icon("fa5s.download", color="#F8F9FA"))
        btn_plantilla.setCursor(Qt.PointingHandCursor)
        btn_plantilla.setToolTip("Descarga un Excel de ejemplo con el formato correcto para importar")
        btn_plantilla.setStyleSheet("QPushButton { background-color: #2D2D2D; color: #F8F9FA; } QPushButton:hover { background-color: #3d3d3d; }")
        btn_plantilla.clicked.connect(self._descargar_plantilla)
        row2.addWidget(btn_plantilla)

        btn_info = QPushButton()
        btn_info.setIcon(qta.icon("fa5s.question-circle", color="#D4AF37"))
        btn_info.setCursor(Qt.PointingHandCursor)
        btn_info.setFixedSize(34, 34)
        btn_info.setToolTip("Ver ayuda sobre las herramientas disponibles")
        btn_info.setStyleSheet("QPushButton { background-color: #2D2D2D; border-radius: 17px; } QPushButton:hover { background-color: #3d3d3d; }")
        btn_info.clicked.connect(self._mostrar_info_importar)
        row2.addWidget(btn_info)

        lista_layout.addLayout(row2)
        self.stack.addWidget(lista_page)

        # Pagina 1: Formulario
        self._form_container = QWidget()
        self._form_layout = QVBoxLayout(self._form_container)
        self._form_layout.setContentsMargins(0, 0, 0, 0)
        self.stack.addWidget(self._form_container)

        layout.addWidget(self.stack)

    def _cargar_lista(self):
        busqueda = self.tabla.input_busqueda.text().strip()
        solo_activos = not self.chk_inactivos.isChecked()
        empleados = empleado_service.listar(busqueda=busqueda, solo_activos=solo_activos)

        orden = self.filtro_orden.currentData() if hasattr(self, 'filtro_orden') else "legajo"
        if orden == "legajo":
            empleados.sort(key=lambda e: int(e.legajo) if e.legajo and e.legajo.isdigit() else 9999)
        elif orden == "nombre":
            empleados.sort(key=lambda e: e.nombre.lower())
        elif orden == "apellido":
            empleados.sort(key=lambda e: (e.apellido or "").lower())

        rows = []
        for e in empleados:
            rows.append((e.id, [
                e.legajo or "",
                e.apellido,
                e.nombre,
                e.dni,
                e.cuil,
                e.departamento.nombre if e.departamento else "",
                e.cargo.nombre if e.cargo else "",
                "Activo" if e.activo else "Inactivo",
            ]))
        self.tabla.set_data(rows)

    def _nuevo(self):
        self._mostrar_form(None)

    def _editar(self, empleado_id: int):
        self._mostrar_form(empleado_id)

    def _editar_seleccionado(self):
        emp_id = self.tabla.selected_id()
        if not emp_id:
            QMessageBox.information(self, "Seleccion", "Selecciona un empleado de la lista.")
            return
        self._editar(emp_id)

    def _ver_detalle(self):
        emp_id = self.tabla.selected_id()
        if not emp_id:
            QMessageBox.information(self, "Seleccion", "Selecciona un empleado de la lista.")
            return
        dialog = EmpleadoDetalleDialog(emp_id, parent=self)
        dialog.editar_clicked.connect(self._editar)
        dialog.exec()

    def _eliminar(self):
        emp_id = self.tabla.selected_id()
        if not emp_id:
            QMessageBox.information(self, "Seleccion", "Selecciona un empleado de la lista.")
            return
        resp = QMessageBox.question(
            self, "Confirmar baja",
            "Estas seguro de dar de baja a este empleado?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp == QMessageBox.Yes:
            empleado_service.eliminar(emp_id)
            self._cargar_lista()

    def _mostrar_form(self, empleado_id: int | None):
        for i in reversed(range(self._form_layout.count())):
            w = self._form_layout.itemAt(i).widget()
            if w:
                w.deleteLater()
        form = FormEmpleado(empleado_id=empleado_id)
        form.guardado.connect(self._on_guardado)
        form.cancelado.connect(self._volver_lista)
        self._form_layout.addWidget(form)
        self.stack.setCurrentIndex(1)

    def _on_guardado(self):
        self._volver_lista()
        self._cargar_lista()

    def _volver_lista(self):
        self.stack.setCurrentIndex(0)

    def _exportar(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from core.config import BASE_DIR

            activos = empleado_service.listar(solo_activos=True)
            activos.sort(key=lambda e: int(e.legajo) if e.legajo and e.legajo.isdigit() else 9999)
            inactivos = empleado_service.listar(solo_activos=False)
            inactivos = [e for e in inactivos if not e.activo]
            inactivos.sort(key=lambda e: int(e.legajo) if e.legajo and e.legajo.isdigit() else 9999)

            wb = Workbook()
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            headers = ["Legajo", "Apellido", "Nombre", label_doc_identidad(), label_id_fiscal(), "Email", "Telefono",
                       "Departamento", "Cargo", "Valor Hora", "Sueldo Mensual", "Tipo Liquidacion"]

            # Hoja Activos
            ws_act = wb.active
            ws_act.title = "Activos"
            for col, h in enumerate(headers, 1):
                cell = ws_act.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for e in activos:
                ws_act.append([
                    e.legajo or "", e.apellido or "", e.nombre or "",
                    e.dni or "", e.cuil or "", e.email or "", e.telefono or "",
                    e.departamento.nombre if e.departamento else "",
                    e.cargo.nombre if e.cargo else "",
                    float(e.valor_hora) if e.valor_hora else 0,
                    float(e.sueldo_mensual) if e.sueldo_mensual else 0,
                    e.tipo_liquidacion or "por_hora",
                ])

            # Hoja Inactivos
            headers_inact = headers + ["Fecha Baja"]
            ws_inact = wb.create_sheet("Inactivos")
            fill_inact = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
            for col, h in enumerate(headers_inact, 1):
                cell = ws_inact.cell(1, col, h)
                cell.font = header_font
                cell.fill = fill_inact
                cell.alignment = Alignment(horizontal="center")
            for e in inactivos:
                fecha_baja = e.updated_at.strftime("%d/%m/%Y") if e.updated_at else ""
                ws_inact.append([
                    e.legajo or "", e.apellido or "", e.nombre or "",
                    e.dni or "", e.cuil or "", e.email or "", e.telefono or "",
                    e.departamento.nombre if e.departamento else "",
                    e.cargo.nombre if e.cargo else "",
                    float(e.valor_hora) if e.valor_hora else 0,
                    float(e.sueldo_mensual) if e.sueldo_mensual else 0,
                    e.tipo_liquidacion or "por_hora",
                    fecha_baja,
                ])

            # Ajustar anchos
            for ws in [ws_act, ws_inact]:
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 16

            output = BASE_DIR / "exports" / "empleados.xlsx"
            output.parent.mkdir(exist_ok=True)
            wb.save(str(output))
            os.startfile(str(output))
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _importar(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar Excel", "", "Excel (*.xlsx *.xls)"
        )
        if not filepath:
            return
        try:
            resultado = importar_empleados(filepath)
            msg = (
                f"Importados: {resultado['importados']}\n"
                f"Duplicados (omitidos): {resultado['duplicados']}\n"
            )
            if resultado["errores"]:
                msg += f"Errores: {len(resultado['errores'])}\n\n"
                msg += "\n".join(resultado["errores"][:10])
            QMessageBox.information(self, "Importacion Completada", msg)
            self._cargar_lista()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al importar: {e}")

    def _descargar_plantilla(self):
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla", "plantilla_empleados.xlsx", "Excel (*.xlsx)"
        )
        if not filepath:
            return
        try:
            generar_plantilla(filepath)
            os.startfile(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _mostrar_info_importar(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Ayuda - Herramientas de Empleados")
        dialog.setFixedSize(720, 680)
        dialog.setModal(True)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        title = QLabel("Herramientas disponibles")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #D4AF37;")
        layout.addWidget(title)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        content = QWidget()
        clayout = QVBoxLayout(content)
        clayout.setSpacing(12)

        # Botones explicados
        clayout.addWidget(self._info_section("Botones", [
            ("Formulario Alta (PDF)", "Genera un PDF en blanco con los campos para que el nuevo empleado complete a mano"),
            ("Exportar Excel", "Exporta el listado completo de empleados (activos e inactivos) a un archivo Excel"),
            ("Importar Excel", "Importa empleados desde un archivo Excel al sistema"),
            ("Descargar Plantilla", "Descarga un Excel de ejemplo con el formato correcto para luego importar"),
        ]))

        # Formato importacion
        lbl_doc = label_doc_identidad()
        lbl_fiscal = label_id_fiscal()
        clayout.addWidget(self._info_section("Formato de Importacion (columnas Excel)", [
            ("Nombre (*)", "Obligatorio. Puede ser nombre completo (se separa automaticamente)"),
            ("Legajo", "Codigo unico. Se genera automaticamente si no se pone"),
            ("Apellido", "Si no se pone, se separa del Nombre"),
            (lbl_doc, "Documento de identidad"),
            (lbl_fiscal, "Identificacion fiscal"),
            ("Email", "Formato email valido"),
            ("Telefono", "Numeros"),
            ("Direccion", "Texto libre"),
            ("Fecha_Nacimiento", "DD/MM/YYYY"),
            ("Fecha_Ingreso", "DD/MM/YYYY. Si no se pone, usa hoy"),
            ("Departamento", "Si no existe, se crea automaticamente"),
            ("Cargo", "Si no existe, se crea automaticamente"),
            ("Valor_Hora", "Numero decimal. Ej: 2500.50"),
            ("Sueldo_Mensual", "Numero decimal. Ej: 450000"),
            ("Tipo_Liquidacion", "por_hora o mensual. Por defecto: por_hora"),
            ("Hora_Entrada", "HH:MM. Ej: 08:00"),
            ("Hora_Salida", "HH:MM. Ej: 17:00"),
            ("Dias_Laborales", "lun,mar,mie,jue,vie (separados por coma)"),
        ]))

        # Notas
        notas = QLabel(
            "Notas:\\n\\n"
            "  - La primera fila del Excel debe tener los nombres de las columnas\\n"
            "  - Si el Legajo ya existe, el registro se omite (no duplica)\\n"
            "  - Departamentos y Cargos nuevos se crean automaticamente"
        )
        notas.setWordWrap(True)
        notas.setStyleSheet("font-size: 12px; padding: 8px;")
        clayout.addWidget(notas)

        scroll.setWidget(content)
        layout.addWidget(scroll)

        btn_cerrar = QPushButton("Entendido")
        btn_cerrar.setMinimumHeight(38)
        btn_cerrar.clicked.connect(dialog.close)
        layout.addWidget(btn_cerrar)

        dialog.exec()

    def _info_section(self, titulo: str, items: list[tuple[str, str]]) -> QFrame:
        frame = QFrame()
        frame.setObjectName("card")
        flayout = QVBoxLayout(frame)
        flayout.setSpacing(6)

        lbl_title = QLabel(titulo)
        lbl_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #D4AF37;")
        flayout.addWidget(lbl_title)

        for campo, desc in items:
            row = QHBoxLayout()
            lbl_campo = QLabel(f"  {campo}")
            lbl_campo.setStyleSheet("font-weight: bold; color: #D4AF37; min-width: 160px;")
            lbl_campo.setFixedWidth(180)
            row.addWidget(lbl_campo)
            lbl_desc = QLabel(desc)
            lbl_desc.setWordWrap(True)
            row.addWidget(lbl_desc)
            row.addStretch()
            flayout.addLayout(row)

        return frame

    def _generar_formulario_alta(self):
        try:
            from services.rrhh.formulario_alta_service import generar_formulario_alta
            filepath = generar_formulario_alta()
            os.startfile(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo generar el formulario: {e}")
